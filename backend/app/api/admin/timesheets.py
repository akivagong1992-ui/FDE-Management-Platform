"""Timesheet API: 单天 / 区间批量创建 + Excel 导入 + 模板下载。

时段语义：
- 1 天最多 3 个时段（上午 / 下午 / 晚上），每段自然 0.5 人天
- 倍率：香港工作日的上下午 1.0×；工作日晚上 + 非工作日全天 1.5×
- 服务端在保存时按 work_date 的 weekday() 决定 is_workday，并算出 weighted_days
"""

import io
from datetime import date as date_cls
from datetime import datetime, timedelta
from decimal import Decimal

from fastapi import (
    APIRouter, Depends, File, HTTPException, Query, Response, UploadFile, status,
)
from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.core.deps import get_current_user, require_role
from app.models.engineer import Engineer
from app.models.project import Project
from app.models.timesheet import (
    SLOT_AFTERNOON, SLOT_EVENING, SLOT_MORNING,
    Timesheet, compute_weighted_days, is_hk_workday,
)
from app.schemas.timesheet import (
    ImportResult,
    ImportRowError,
    TimesheetCreate,
    TimesheetOut,
    TimesheetRangeCreate,
    TimesheetRangeResult,
    TimesheetReject,
    TimesheetUpdate,
)

router = APIRouter(prefix="/timesheets", tags=["timesheets"])

# 管理者角色（能审批 / 看全部）
PM_ROLES = {"admin", "lead", "pm", "finance"}
ENGINEER_ROLE = "engineer"


def _to_out(t: Timesheet) -> TimesheetOut:
    return TimesheetOut(
        id=t.id,
        engineer_id=t.engineer_id,
        engineer_name=t.engineer.full_name if t.engineer else None,
        project_id=t.project_id,
        project_name=t.project.name if t.project else None,
        assignment_id=t.assignment_id,
        work_date=t.work_date,
        has_morning=t.has_morning,
        has_afternoon=t.has_afternoon,
        has_evening=t.has_evening,
        is_workday=t.is_workday,
        natural_days=t.natural_days,
        weighted_days=t.weighted_days,
        description=t.description,
        approval_status=t.approval_status,
        reject_reason=t.reject_reason,
        reviewed_at=t.reviewed_at,
        submitted_by_user_id=t.submitted_by_user_id,
        is_approved=t.is_approved,
        created_at=t.created_at,
    )


@router.get("", response_model=list[TimesheetOut])
async def list_timesheets(
    engineer_id: int | None = None,
    project_id: int | None = None,
    date_from: str | None = Query(default=None),
    date_to: str | None = Query(default=None),
    approval_filter: str | None = Query(default=None),
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(get_current_user),
) -> list[TimesheetOut]:
    stmt = select(Timesheet).order_by(Timesheet.work_date.desc(), Timesheet.id.desc())

    # engineer 角色：强制只看自己的工时记录
    if user.get("role") == ENGINEER_ROLE:
        eng_id = user.get("engineer_id")
        if not eng_id:
            return []
        stmt = stmt.where(Timesheet.engineer_id == eng_id)
    elif user.get("role") not in PM_ROLES:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Forbidden")

    if engineer_id is not None:
        stmt = stmt.where(Timesheet.engineer_id == engineer_id)
    if project_id is not None:
        stmt = stmt.where(Timesheet.project_id == project_id)
    if approval_filter:
        stmt = stmt.where(Timesheet.approval_status == approval_filter)
    if date_from:
        stmt = stmt.where(Timesheet.work_date >= date_cls.fromisoformat(date_from))
    if date_to:
        stmt = stmt.where(Timesheet.work_date <= date_cls.fromisoformat(date_to))
    rows = (await db.execute(stmt)).scalars().all()
    return [_to_out(t) for t in rows]


async def _build_and_save_one(
    db: AsyncSession, payload: TimesheetCreate,
    *, commit: bool = True, submitted_by: int | None = None,
) -> Timesheet:
    if not await db.get(Engineer, payload.engineer_id):
        raise HTTPException(status_code=400, detail=f"工程师 #{payload.engineer_id} 不存在")
    if not await db.get(Project, payload.project_id):
        raise HTTPException(status_code=400, detail=f"项目 #{payload.project_id} 不存在")
    is_wd = is_hk_workday(payload.work_date)
    natural, weighted = compute_weighted_days(
        payload.work_date, payload.has_morning, payload.has_afternoon, payload.has_evening,
        is_workday=is_wd,
    )
    t = Timesheet(
        engineer_id=payload.engineer_id, project_id=payload.project_id,
        assignment_id=payload.assignment_id, work_date=payload.work_date,
        has_morning=payload.has_morning, has_afternoon=payload.has_afternoon,
        has_evening=payload.has_evening, is_workday=is_wd,
        natural_days=natural, weighted_days=weighted,
        description=payload.description,
        approval_status="pending",
        submitted_by_user_id=submitted_by,
    )
    db.add(t)
    try:
        if commit:
            await db.commit()
            await db.refresh(t)
    except IntegrityError as e:
        await db.rollback()
        raise HTTPException(
            status_code=400,
            detail=f"{payload.work_date} 当天该工程师在该项目的工时已存在；如需补时段请编辑",
        ) from e
    return t


def _force_engineer_id_if_engineer(payload, user: dict) -> None:
    """engineer 角色不能替别人提交工时，强制改为自己的 engineer_id。"""
    if user.get("role") == ENGINEER_ROLE:
        own = user.get("engineer_id")
        if not own:
            raise HTTPException(status_code=403, detail="当前账号未绑定工程师记录")
        payload.engineer_id = own


@router.post("", response_model=TimesheetOut, status_code=status.HTTP_201_CREATED)
async def create_timesheet(
    payload: TimesheetCreate,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(get_current_user),
) -> TimesheetOut:
    if user.get("role") not in PM_ROLES and user.get("role") != ENGINEER_ROLE:
        raise HTTPException(status_code=403, detail="Forbidden")
    _force_engineer_id_if_engineer(payload, user)
    t = await _build_and_save_one(db, payload, submitted_by=user.get("user_id"))
    return _to_out(t)


@router.post("/range", response_model=TimesheetRangeResult, status_code=status.HTTP_201_CREATED)
async def create_timesheet_range(
    payload: TimesheetRangeCreate,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(get_current_user),
) -> TimesheetRangeResult:
    if user.get("role") not in PM_ROLES and user.get("role") != ENGINEER_ROLE:
        raise HTTPException(status_code=403, detail="Forbidden")
    _force_engineer_id_if_engineer(payload, user)
    """起止日期 + 时段集合 → 逐日展开，每天一条记录。失败的日期记入 skipped。"""
    has_morning = SLOT_MORNING in payload.slots
    has_afternoon = SLOT_AFTERNOON in payload.slots
    has_evening = SLOT_EVENING in payload.slots

    created: list[TimesheetOut] = []
    skipped: list[ImportRowError] = []
    total_natural = Decimal("0")
    total_weighted = Decimal("0")

    cur = payload.start_date
    idx = 0
    while cur <= payload.end_date:
        idx += 1
        single = TimesheetCreate(
            engineer_id=payload.engineer_id, project_id=payload.project_id,
            assignment_id=payload.assignment_id, work_date=cur,
            has_morning=has_morning, has_afternoon=has_afternoon, has_evening=has_evening,
            description=payload.description,
        )
        try:
            t = await _build_and_save_one(db, single, submitted_by=user.get("user_id"))
            out = _to_out(t)
            created.append(out)
            total_natural += out.natural_days
            total_weighted += out.weighted_days
        except HTTPException as he:
            skipped.append(ImportRowError(row=idx, message=f"{cur}: {he.detail}"))
        cur += timedelta(days=1)

    return TimesheetRangeResult(
        created=created, skipped=skipped,
        total_natural_days=total_natural, total_weighted_days=total_weighted,
    )


@router.patch("/{ts_id}", response_model=TimesheetOut)
async def update_timesheet(
    ts_id: int,
    payload: TimesheetUpdate,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(get_current_user),
) -> TimesheetOut:
    t = await db.get(Timesheet, ts_id)
    if not t:
        raise HTTPException(status_code=404, detail="工时记录不存在")
    # engineer 只能改自己的、且非已审；改完状态回 pending（重审）
    if user.get("role") == ENGINEER_ROLE:
        if t.engineer_id != user.get("engineer_id"):
            raise HTTPException(status_code=403, detail="无权编辑他人的工时")
        if t.approval_status == "approved":
            raise HTTPException(status_code=400, detail="已审通过的工时不能改；如需修改请联系管理者")
        t.approval_status = "pending"
        t.reject_reason = None
        t.reviewed_at = None
        t.is_approved = False
    elif user.get("role") not in PM_ROLES:
        raise HTTPException(status_code=403, detail="Forbidden")

    for k, v in payload.model_dump(exclude_unset=True).items():
        setattr(t, k, v)
    t.recompute()
    await db.commit()
    await db.refresh(t)
    return _to_out(t)


@router.delete("/{ts_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_timesheet(
    ts_id: int,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(get_current_user),
) -> None:
    t = await db.get(Timesheet, ts_id)
    if not t:
        raise HTTPException(status_code=404, detail="工时记录不存在")
    # engineer 只能删自己的、且非已审通过
    if user.get("role") == ENGINEER_ROLE:
        if t.engineer_id != user.get("engineer_id"):
            raise HTTPException(status_code=403, detail="无权删除他人的工时")
        if t.approval_status == "approved":
            raise HTTPException(status_code=400, detail="已审通过的工时不能删；如需撤回请联系管理者")
    elif user.get("role") not in PM_ROLES:
        raise HTTPException(status_code=403, detail="Forbidden")
    await db.delete(t)
    await db.commit()


@router.patch("/{ts_id}/approve", response_model=TimesheetOut)
async def approve_timesheet(
    ts_id: int,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(require_role("admin", "lead", "finance")),
) -> TimesheetOut:
    t = await db.get(Timesheet, ts_id)
    if not t:
        raise HTTPException(status_code=404, detail="工时记录不存在")
    t.approval_status = "approved"
    t.is_approved = True
    t.reject_reason = None
    t.reviewed_at = datetime.utcnow()
    t.reviewed_by_user_id = user.get("user_id")
    await db.commit()
    await db.refresh(t)
    return _to_out(t)


@router.patch("/{ts_id}/reject", response_model=TimesheetOut)
async def reject_timesheet(
    ts_id: int,
    payload: TimesheetReject,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(require_role("admin", "lead", "finance")),
) -> TimesheetOut:
    t = await db.get(Timesheet, ts_id)
    if not t:
        raise HTTPException(status_code=404, detail="工时记录不存在")
    t.approval_status = "rejected"
    t.is_approved = False
    t.reject_reason = payload.reason.strip()
    t.reviewed_at = datetime.utcnow()
    t.reviewed_by_user_id = user.get("user_id")
    await db.commit()
    await db.refresh(t)
    return _to_out(t)


# ─── Excel template + import ─────────────────────────────────────────

EXCEL_HEADERS = [
    "工程师姓名", "项目编号或名称", "工作日期(YYYY-MM-DD)",
    "上午(0/1)", "下午(0/1)", "晚上(0/1)", "描述(可选)",
]


def _to_bool(v) -> bool:
    if v is None or v == "":
        return False
    s = str(v).strip().lower()
    return s in ("1", "true", "yes", "y", "是", "✓", "x")


@router.get("/template")
async def download_template(
    _: dict = Depends(get_current_user),
) -> Response:
    wb = Workbook()
    ws = wb.active
    ws.title = "工时导入"
    ws.append(EXCEL_HEADERS)
    ws.append(["李志强", "PROJ-001", "2026-05-23", 1, 1, 0, "中环现场调试整天"])
    ws.append(["李志强", "PROJ-001", "2026-05-24", 0, 0, 1, "周末晚上加班"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="timesheet-template.xlsx"'},
    )


@router.post("/import-excel", response_model=ImportResult)
async def import_excel(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    _: dict = Depends(require_role("admin", "lead", "finance")),
) -> ImportResult:
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx / .xls 文件")
    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel 解析失败：{e}") from e
    ws = wb.active

    rows_iter = ws.iter_rows(min_row=2, values_only=True)
    created, skipped, errors = 0, 0, []
    eng_by_name: dict[str, Engineer] = {}
    proj_by_key: dict[str, Project] = {}

    for idx, row in enumerate(rows_iter, start=2):
        try:
            if not row or all(c is None for c in row):
                continue
            (eng_name_raw, proj_key_raw, date_raw,
             am_raw, pm_raw, eve_raw, desc_raw) = (list(row) + [None] * 7)[:7]
            if eng_name_raw is None or proj_key_raw is None or date_raw is None:
                errors.append(ImportRowError(row=idx, message="缺少必填字段（工程师/项目/日期）"))
                skipped += 1
                continue

            eng_name = str(eng_name_raw).strip()
            proj_key = str(proj_key_raw).strip()

            if eng_name not in eng_by_name:
                e = (await db.execute(
                    select(Engineer).where(Engineer.full_name == eng_name)
                )).scalar_one_or_none()
                if not e:
                    errors.append(ImportRowError(row=idx, message=f"工程师 '{eng_name}' 不存在"))
                    skipped += 1
                    continue
                eng_by_name[eng_name] = e
            engineer = eng_by_name[eng_name]

            if proj_key not in proj_by_key:
                p = (await db.execute(
                    select(Project).where(Project.code == proj_key)
                )).scalar_one_or_none()
                if not p:
                    p = (await db.execute(
                        select(Project).where(Project.name == proj_key)
                    )).scalar_one_or_none()
                if not p:
                    errors.append(ImportRowError(row=idx, message=f"项目 '{proj_key}' 不存在"))
                    skipped += 1
                    continue
                proj_by_key[proj_key] = p
            project = proj_by_key[proj_key]

            if isinstance(date_raw, datetime):
                work_date = date_raw.date()
            elif isinstance(date_raw, date_cls):
                work_date = date_raw
            else:
                work_date = date_cls.fromisoformat(str(date_raw).strip())

            am = _to_bool(am_raw); pm = _to_bool(pm_raw); eve = _to_bool(eve_raw)
            if not (am or pm or eve):
                errors.append(ImportRowError(row=idx, message="上午/下午/晚上 至少选一个"))
                skipped += 1
                continue

            payload = TimesheetCreate(
                engineer_id=engineer.id, project_id=project.id,
                work_date=work_date, has_morning=am, has_afternoon=pm, has_evening=eve,
                description=str(desc_raw).strip() if desc_raw else None,
            )
            try:
                await _build_and_save_one(db, payload)
                created += 1
            except HTTPException as he:
                errors.append(ImportRowError(row=idx, message=str(he.detail)))
                skipped += 1
        except Exception as e:
            errors.append(ImportRowError(row=idx, message=str(e)))
            skipped += 1

    return ImportResult(created=created, skipped=skipped, errors=errors)
